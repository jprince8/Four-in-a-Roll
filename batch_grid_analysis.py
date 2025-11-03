#!/usr/bin/env python3
"""
Batch wrapper around four_in_a_roll_analysis.render_full_logic.

Generates solved outcomes for a range of board dimensions, running each
analysis in parallel and writing the first-node outcome (P1 win, P2 win,
ongoing) into a colour-coded spreadsheet.
"""

import argparse
import concurrent.futures
import importlib
import itertools
import json
import os
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple, cast

from tqdm import tqdm


DEFAULT_MIN_ROWS = 4
DEFAULT_MAX_ROWS = 7
DEFAULT_MIN_COLS = 4
DEFAULT_MAX_COLS = 7
MAX_WORKERS = None  # Use os.cpu_count() by default


@dataclass(frozen=True)
class BoardSpec:
    rows: int
    cols: int


@dataclass(frozen=True)
class RenderOptions:
    plot_pdf: bool
    exclude_wins: bool


@dataclass
class BoardResult:
    spec: BoardSpec
    turns: int
    max_game_length: int
    outcome_label: str
    status: str
    fill_colour: str
    json_path: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Batch Four-in-a-Roll analysis across multiple board sizes."
    )
    parser.add_argument(
        "--min-rows",
        type=int,
        default=DEFAULT_MIN_ROWS,
        help=f"minimum row count (inclusive, default: {DEFAULT_MIN_ROWS})",
    )
    parser.add_argument(
        "--max-rows",
        type=int,
        default=DEFAULT_MAX_ROWS,
        help=f"maximum row count (inclusive, default: {DEFAULT_MAX_ROWS})",
    )
    parser.add_argument(
        "--min-cols",
        type=int,
        default=DEFAULT_MIN_COLS,
        help=f"minimum column count (inclusive, default: {DEFAULT_MIN_COLS})",
    )
    parser.add_argument(
        "--max-cols",
        type=int,
        default=DEFAULT_MAX_COLS,
        help=f"maximum column count (inclusive, default: {DEFAULT_MAX_COLS})",
    )
    return parser.parse_args()


def build_jobs(args: argparse.Namespace) -> List[BoardSpec]:
    if args.min_rows > args.max_rows:
        raise ValueError("min-rows must be ≤ max-rows")
    if args.min_cols > args.max_cols:
        raise ValueError("min-cols must be ≤ max-cols")
    if args.min_rows < 1 or args.min_cols < 1:
        raise ValueError("Row/column counts must be positive.")

    rows = range(args.min_rows, args.max_rows + 1)
    cols = range(args.min_cols, args.max_cols + 1)
    specs = [BoardSpec(r, c) for r, c in itertools.product(rows, cols)]
    return specs


def _summarise_root(node: Dict[str, object], p1_token: int, p2_token: int) -> Tuple[str, str]:
    forced_winner = node.get("forced_winner")
    status = str(node.get("status", ""))
    status_lower = status.lower()
    if forced_winner == p1_token:
        return "P1 win", "FFFF5C5C"  # red fill
    if forced_winner == p2_token:
        return "P2 win", "FFFFEE58"  # yellow fill
    if "ongoing" in status_lower:
        return "Draw", "FFB0BEC5"  # grey fill
    return "Draw", "FFB0BEC5"


def analyse_board(spec: BoardSpec) -> BoardResult:
    # Local import for isolated globals per process.
    analysis_module = importlib.import_module("four_in_a_roll_analysis")
    analysis = cast(Any, analysis_module)

    turns = spec.rows * spec.cols

    analysis.INIT_ROWS = spec.rows
    analysis.INIT_COLS = spec.cols
    json_path, pdf_path = analysis.render_full_logic(
        turns,
        disable_progress=True,
    )

    with open(json_path, "r") as handle:
        payload = json.load(handle)

    nodes_list = payload.get("nodes", [])
    if not nodes_list:
        raise RuntimeError(f"No nodes found in analysis output for board {spec.rows}x{spec.cols}")

    max_game_length = max(int(node.get("layer", 0) or 0) for node in nodes_list)

    nodes = {node["id"]: node for node in nodes_list if "id" in node}
    root = nodes.get(0)
    if root is None:
        root = nodes_list[0]
    if root is None:
        raise RuntimeError(f"Unable to locate root node for board {spec.rows}x{spec.cols}")

    outcome_label, fill = _summarise_root(root, analysis.P1, analysis.P2)
    status = str(root.get("status", ""))
    return BoardResult(
        spec=spec,
        turns=turns,
        max_game_length=max_game_length,
        outcome_label=outcome_label,
        status=status,
        fill_colour=fill,
        json_path=json_path,
    )


def collect_results(
    specs: Iterable[BoardSpec],
) -> List[BoardResult]:
    specs_list = list(specs)
    results: List[BoardResult] = []
    with concurrent.futures.ProcessPoolExecutor(max_workers=MAX_WORKERS) as pool, tqdm(
        total=len(specs_list),
        desc="Boards",
        unit="board",
    ) as progress:
        future_to_spec = {
            pool.submit(analyse_board, spec): spec for spec in specs_list
        }
        for future in concurrent.futures.as_completed(future_to_spec):
            spec = future_to_spec[future]
            try:
                result = future.result()
            except Exception as exc:
                raise RuntimeError(f"Analysis failed for board {spec.rows}x{spec.cols}") from exc
            results.append(result)
            progress.update(1)
            progress.write(f"[done] {spec.rows}x{spec.cols}: {result.outcome_label} (max {result.max_game_length}/{result.turns} turns)")
    results.sort(key=lambda r: (r.spec.rows, r.spec.cols))
    return results


def write_spreadsheet(
    results: List[BoardResult],
    spreadsheet_path: str,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to write the spreadsheet. Install it with "
            "`pip install openpyxl`."
        ) from exc

    rows_sorted = sorted({r.spec.rows for r in results})
    cols_sorted = sorted({r.spec.cols for r in results})

    lookup: Dict[Tuple[int, int], BoardResult] = {
        (r.spec.rows, r.spec.cols): r for r in results
    }

    wb = Workbook()
    ws = cast(Any, wb.active)
    ws.title = "Outcomes"

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value="Rows \\ Cols").font = header_font

    for idx, col in enumerate(cols_sorted, start=2):
        cell = ws.cell(row=1, column=idx, value=col)
        cell.font = header_font
        cell.alignment = center

    for idx, row in enumerate(rows_sorted, start=2):
        cell = ws.cell(row=idx, column=1, value=row)
        cell.font = header_font
        cell.alignment = center

    for r_idx, row in enumerate(rows_sorted, start=2):
        for c_idx, col in enumerate(cols_sorted, start=2):
            result = lookup.get((row, col))
            cell = ws.cell(row=r_idx, column=c_idx)
            if result is None:
                cell.value = "n/a"
                continue
            cell.value = f"{result.outcome_label} (max {result.max_game_length}/{result.turns})"
            cell.alignment = center
            cell.fill = PatternFill(start_color=result.fill_colour, end_color=result.fill_colour, fill_type="solid")

    for column_cells in ws.columns:
        column_cells = list(column_cells)
        if not column_cells:
            continue
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        adjusted = max(10, max_length + 2)
        first_cell = column_cells[0]
        column_letter = cast(Any, first_cell).column_letter
        dimension = ws.column_dimensions[column_letter]
        cast(Any, dimension).width = adjusted

    os.makedirs(os.path.dirname(spreadsheet_path) or ".", exist_ok=True)
    wb.save(spreadsheet_path)
    print(f"Wrote spreadsheet → {spreadsheet_path}")


def main() -> None:
    args = parse_args()
    specs = build_jobs(args)
    spreadsheet_path = os.path.join(
        "outputs",
        f"batch_summary_r{args.min_rows}-{args.max_rows}_c{args.min_cols}-{args.max_cols}.xlsx",
    )
    print(
        f"Running analysis for {len(specs)} boards: "
        f"rows {args.min_rows}-{args.max_rows}, cols {args.min_cols}-{args.max_cols}"
    )
    results = collect_results(specs)
    write_spreadsheet(results, spreadsheet_path)

    print("\nSummary:")
    for result in results:
        print(
            f"  {result.spec.rows}x{result.spec.cols} "
            f"→ {result.outcome_label} (max {result.max_game_length}/{result.turns} turns until win)"
        )


if __name__ == "__main__":
    main()
